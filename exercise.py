# ITP Week 3 Day 3 Exercise
from openpyxl import Workbook

import requests

import time

character_url = "https://rickandmortyapi.com/api/character"

# RICK AND MORTY API DOCS: https://rickandmortyapi.com/documentation

# we want to make a copy of the Rick and Morty database (which is provided through the api)

wb = Workbook()
ws = wb.active
ws.title = "Rick and Morty Characters"

response = requests.get(character_url)
data = response.json()
# EASY MODE

# import the appropriate modules (you have 3)

next_url = character_url
all_characters = []

while next_url:
    next_response = requests.get(next_url)
    next_data = next_response.json()
    all_characters.extend(next_data["results"])
    next_url = next_data["info"]["next"]
    print(len(all_characters))
    time.sleep(0.5)


# set up a workbook and worksheet titled "Rick and Morty Characters"

print(data["results"][0].keys())

headers = ["id", "name", "status", "species", "type", "gender", "origin", "location", "image", "episode", "url", "created"]
# # assign a variable 'data' with the returned GET request
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)
# create the appropriate headers in openpyxl for all of the keys for a single character
for row_num, character in enumerate(all_characters, 2):
    if character["episode"]: 
        print("processing episodes for:", character["name"])
        current_episode = []
        for episode_url in character["episode"]:
            try:
                ep_response = requests.get(episode_url)
                ep_data = ep_response.json()
                current_episode.append(ep_data["name"])
                time.sleep(0.3)
            except:
                current_episode.append(episode_url)
        character["episode"] = current_episode
    for col_num, value in enumerate(character.values(), 1):
        ws.cell(row=row_num, column=col_num, value=str(value))

# loop through all of the 'results' of the data to populate the rows and columns for each character

# NOTE: due to the headers, the rows need to be offset by one!

# MEDIUM MODE

# create 2 new worksheets for "Rick and Morty Locations" and "Rick and Morty Episodes"

# create 2 new variables for episode_url and location_url (retrieve it from the docs!)

location_url = "https://rickandmortyapi.com/api/location"

location_response = requests.get(location_url)
location_data = location_response.json()

# populate the new worksheets appropriately with all of the data!
print(location_data["results"][0].keys())

next_url = location_url
all_location = []


while next_url:
    next_response = requests.get(next_url)
    next_data = next_response.json()
    all_location.extend(next_data["results"])
    next_url = next_data["info"]["next"]
    print(len(all_location))
    time.sleep(0.5)


ws2 = wb.create_sheet("Rick and Morty Location")
# NOTE: don't forget your headers!

headers2 = ["id", "name", "type", "dimension", "residents", "url", "created"]
for col_num, header in enumerate(headers2, 1):
    ws2.cell(row=1, column=col_num, value=header)
for row_num, location in enumerate(all_location, 2):
    for col_num, value in enumerate(location.values(), 1):
        ws2.cell(row=row_num, column=col_num, value=str(value)) 

episode_url = "https://rickandmortyapi.com/api/episode"

episode_response = requests.get(episode_url)
episode_data = episode_response.json()

print(episode_data["results"][0].keys())

next_url = episode_url
all_episode = []


while next_url:
    next_response = requests.get(next_url)
    next_data = next_response.json()
    all_episode.extend(next_data["results"])
    next_url = next_data["info"]["next"]
    print(len(all_episode))
    time.sleep(0.5)

ws3 = wb.create_sheet("Rick and Morty Episode")

headers3 =['id', 'name', 'air_date', 'episode', 'characters', 'url', 'created']

for col_num, header in enumerate(headers3, 1):
    ws3.cell(row=1, column=col_num, value=header)
for row_num, episode in enumerate(all_episode, 2):
    for col_num, value in enumerate(episode.values(), 1):
        ws3.cell(row=row_num, column=col_num, value=str(value)) 

# HARD MODE
# Can you decipher the INFO key of the data to use "next" url to continuously pull data?
# Currently, we are only pulling 20 items per api pull!
# WE WANT EVERYTHING. (contact instructors for office hours if stuck!)

print(data["info"])

    
    
# NIGHTMARE
# The inner information for characters, locations, and episodes, references one another through urls
# ie. for episode 28, it lists all the character but by their url
# can you use the URLs to make a subsequent call inside your for loops
# to replace the url with just the appropriate names?
# NOTE: need to make use of if statements to see if url exists or not
# (contact instructors for office hours if stuck!)


wb.save("./spreadsheets/exercise.xlsx")
